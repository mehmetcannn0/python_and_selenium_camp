import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
from constants import globalConstants

class Test_Sauce:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        # self.driver.maximize_window()
        self.driver.minimize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = "test_gün_6_ödev_2/"+str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()
    def get_data_for_invalid_login():
        #veriyi al
        excelFile = openpyxl.load_workbook(globalConstants.name_of_excel)
        selectedSheet = excelFile["invalid_login"]
        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
    
    def get_data_for_product_and_price():
        excelFile = openpyxl.load_workbook(globalConstants.name_of_excel)
        selectedSheet = excelFile["product_and_price"]
        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            name = str(selectedSheet.cell(i,1).value)
            price = str(selectedSheet.cell(i,2).value)
            tupleData = (name,price)
            data.append(tupleData)
        return data
    
    def get_data_for_add_to_cart_button():
        excelFile = openpyxl.load_workbook(globalConstants.name_of_excel)
        selectedSheet = excelFile["add_to_cart_button"]
        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            index = selectedSheet.cell(i,1).value        
            tupleData = (index)
            data.append(tupleData)
        return data
    
    def get_data_for_add_to_cart_process():
        excelFile = openpyxl.load_workbook(globalConstants.name_of_excel)
        selectedSheet = excelFile["add_to_cart_process"]
        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            count_of_product_to_add_to_cart = selectedSheet.cell(i,1).value        
            tupleData = (count_of_product_to_add_to_cart)
            data.append(tupleData)
        return data

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    @pytest.mark.parametrize("username,password",get_data_for_invalid_login())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login-{username}-{password}.png")
        assert errorMessage.text == globalConstants.invalid_login_error_message_text

    def test_user_name_req(self):
        self.waitForElementVisible((By.ID,"user-name"))
        self.waitForElementVisible((By.ID,"password"),10)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_user_name_req-__-__.png")
        assert errorMessage.text == globalConstants.user_name_req_errorMessage_text


    def test_pass_req(self):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        usernameInput.send_keys("1")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_pass_req-1-__.png")
        assert errorMessage.text == globalConstants.pass_req_errorMessage_text
      

    def test_locked_out_user(self):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("locked_out_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_locked_out_user-locked_out_user-secret_sauce.png")
        assert errorMessage.text == globalConstants.locked_out_user_errorMessage_text
  

    def test_x_icon(self):
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        sleep(2)
        testResult=True
        
        # ^^ suan 2 adet x olmalı    
        x_icons = self.driver.find_elements(By.CLASS_NAME,"error_icon")
        num_of_x_icons=len(x_icons)

        
        if(num_of_x_icons==2):
            testResult=True

            error_button=self.driver.find_element(By.CLASS_NAME,"error-button")
            error_button.click()
            x_icons.clear()
            sleep(2)
            x_icons = self.driver.find_elements(By.CLASS_NAME,"error_icon")
            num_of_x_icons=len(x_icons)

            if(num_of_x_icons==0):
                # ^^ suan hiç x olmamalı
      
                print("x_icon fonksiyonu")
                self.driver.save_screenshot(f"{self.folderPath}/test_x_icon-__-__.png")
                assert testResult          

            else:
                testResult=False 

                print(f"error-button'a tıklandıktan sonra {num_of_x_icons} X ikonu bulundu \nHiç bulunmamalıydı")   
                print("x_icon fonksiyonu")
                self.driver.save_screenshot(f"{self.folderPath}/test_x_icon-__-__.png")
                assert testResult
                               
        else:
            testResult=False

            print(f"{num_of_x_icons} adet X ikonu bulundu \n2 adet olmalıydı")
            print("x_icon fonksiyonu")
            self.driver.save_screenshot(f"{self.folderPath}/test_x_icon-__-__.png")
            assert testResult      

    def test_standard_user(self):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        current_url =self.driver.current_url

        if(current_url.endswith("/inventory.html")):
            testResult=True

            print("standard_user fonksiyonu")
            self.driver.save_screenshot(f"{self.folderPath}/test_standard_user-standard_user-secret_sauce.png")
            assert testResult

        else:
            testResult=False

            print("standard_user fonksiyonu")
            self.driver.save_screenshot(f"{self.folderPath}/test_standard_user-standard_user-secret_sauce.png")
            assert testResult

    def test_products(self):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        
        self.waitForElementVisible((By.CLASS_NAME,"inventory_item"))
        products=self.driver.find_elements(By.CLASS_NAME,"inventory_item")
        num_of_products=len(products)
        if(num_of_products==6):
            testResult=True

            print("products fonksiyonu")
            self.driver.save_screenshot(f"{self.folderPath}/test_products-standard_user-secret_sauce.png")
            assert testResult
        else:
            testResult=False

            print(f"{num_of_products} adet ürün bulundu \n6 adet bulunmalıydı") 
            print("products fonksiyonu")
            self.driver.save_screenshot(f"{self.folderPath}/test_products-standard_user-secret_sauce.png")
            assert testResult

       
    @pytest.mark.parametrize("name,price",get_data_for_product_and_price())
    def test_product_and_price(self,name,price):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.waitForElementVisible((By.CLASS_NAME,"inventory_item_name"))
        name_of_products_elements=self.driver.find_elements(By.CLASS_NAME,"inventory_item_name")
        price_of_products_elements=self.driver.find_elements(By.CLASS_NAME,"inventory_item_price")
        name_of_products=[]
        price_of_products=[]

        for n in name_of_products_elements:
            name_of_products.append(n.text)
        for p in price_of_products_elements:
            price_of_products.append(p.text)

        if(name_of_products.count(name)>0 and price_of_products.count(price)>0):
            index_of_name=name_of_products.index(name)
            index_of_price=price_of_products.index(price)
            if(index_of_name==index_of_price):
                print("test_product_and_price fonksiyonu")
                self.driver.save_screenshot(f"{self.folderPath}/test_product_and_price-{name}-{price}.png")
                assert index_of_name==index_of_price
            else:
                print("Ürün ile fiyat eşleşmiyor") 
                print("test_product_and_price fonksiyonu")
                self.driver.save_screenshot(f"{self.folderPath}/test_product_and_price-{name}-{price}.png")
                assert index_of_name==index_of_price
        else:    
            print("Ürün yada fiyat bulunamadı") 
            print("test_product_and_price fonksiyonu")
            self.driver.save_screenshot(f"{self.folderPath}/test_product_and_price-{name}-{price}.png")
            assert name_of_products.count(name)>0 and price_of_products.count(price)>0


    @pytest.mark.parametrize("index",get_data_for_add_to_cart_button())
    def test_add_to_cart_button(self,index):
        self.waitForElementVisible((By.ID,"user-name"))        
        usernameInput = self.driver.find_element(By.ID, "user-name")
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        
        self.waitForElementVisible((By.CLASS_NAME,"btn_inventory"))
        add_and_remove_card_btns=self.driver.find_elements(By.CLASS_NAME,"btn_inventory")

        if(len(add_and_remove_card_btns)>=index and add_and_remove_card_btns[index].text=="Add to cart"):

            add_and_remove_card_btns[index].click()            
            add_and_remove_card_btns=[]
            add_and_remove_card_btns=self.driver.find_elements(By.CLASS_NAME,"btn_inventory")
            assert add_and_remove_card_btns[index].text=="Remove"


    @pytest.mark.parametrize("count_of_product_to_add_to_cart",get_data_for_add_to_cart_process())
    def test_add_to_cart_process(self,count_of_product_to_add_to_cart):
        self.waitForElementVisible((By.ID,"user-name"))
        
        usernameInput = self.driver.find_element(By.ID, "user-name")
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        
        self.waitForElementVisible((By.CLASS_NAME,"btn_inventory"))
        add_and_remove_card_btns=self.driver.find_elements(By.CLASS_NAME,"btn_inventory")

        for index in range(count_of_product_to_add_to_cart):
            if(add_and_remove_card_btns[index].text=="Add to cart"):
                add_and_remove_card_btns[index].click()            
        self.driver.save_screenshot(f"{self.folderPath}/test_add_to_cart_process-{count_of_product_to_add_to_cart}.png")
        assert self.driver.find_element(By.CLASS_NAME,"shopping_cart_badge").text==str(count_of_product_to_add_to_cart)

    def test_logout(self):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()

        self.waitForElementVisible((By.ID,"react-burger-menu-btn"),10)
        menu = self.driver.find_element(By.ID,"react-burger-menu-btn")

        menu.click()
        self.waitForElementVisible((By.ID,"logout_sidebar_link"),10)
        logoutBtn = self.driver.find_element(By.ID,"logout_sidebar_link")

        logoutBtn.click()
        WebDriverWait(self.driver,5).until(ec.url_to_be(globalConstants.URL))

        print(self.driver.current_url)

        self.driver.save_screenshot(f"{self.folderPath}/test_logout.png")
        assert ec.url_to_be(globalConstants.URL)